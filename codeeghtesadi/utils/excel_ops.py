import os
import glob
import time
import math
import numpy as np
import pandas as pd
import xlwings as xw
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, fills
from openpyxl import styles
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import openpyxl.utils as utils
from concurrent.futures import ThreadPoolExecutor, wait

# Internal imports
from codeeghtesadi.utils.decorators import log_the_func, measure_time, check_if_up_to_date
from codeeghtesadi.utils.common import maybe_make_dir, get_filename, get_update_date
from codeeghtesadi.utils.database import drop_into_db, connect_to_sql, n_retries # n_retries might be needed or handled in db
# n_retries is global in helpers, locally used in connect_to_sql. Here we don't need it unless we reimplement retry logic.

@measure_time
def read_multiple_excel_sheets(path, sheet_name=None):
    """
    Read all sheets from the Excel file into a single concatenated DataFrame.
    """
    df_dict = pd.read_excel(path, sheet_name=sheet_name)
    
    if isinstance(df_dict, pd.DataFrame):
        return df_dict

    # Concatenate all DataFrames in the dictionary into a single DataFrame
    return pd.concat(df_dict.values(), ignore_index=True)

@log_the_func('none')
def remove_excel_files(files=None, pathsave=None, file_path=None, postfix=['html'], *args, **kwargs):
    """
    Remove specified files or files with certain extensions from a directory.
    """
    if files is None:
        files = []
    else:
        # Avoid modifying the original list if it's passed as a default arg (though here it's None by default)
        files = list(files)

    if file_path is not None:
        for item in postfix:
            # Add dot if missing
            pattern = item if item.startswith('.') else f".{item}"
            files.extend(glob.glob(os.path.join(file_path, f"*{pattern}")))

    for f in files:
        if os.path.exists(f):
            try:
                os.remove(f)
            except Exception as e:
                print(f"Error removing file {f}: {e}")

def _drop_data_to_db(df, table_name, append_to_prev=False, db_name=None, multithread=False, num_chunks=None):
    """
    Internal helper to drop a DataFrame into the database with optional chunking/multithreading.
    """
    if df.empty:
        return

    columns = df.columns.tolist()
    
    if multithread:
        n_workers = num_chunks or 10
        chunks = np.array_split(df, n_workers)
        
        # First chunk to handle table creation/truncation
        drop_into_db(table_name=table_name,
                     columns=columns,
                     values=chunks[0].values.tolist(),
                     append_to_prev=append_to_prev,
                     db_name=db_name)
        
        if len(chunks) > 1:
            with ThreadPoolExecutor(max_workers=n_workers-1) as executor:
                jobs = [executor.submit(drop_into_db, table_name,
                                        columns,
                                        chunk.values.tolist(),
                                        True, # Always append for subsequent chunks
                                        db_name)
                        for chunk in chunks[1:]]
                wait(jobs)
    else:
        # If very large, chunk it sequentially to avoid memory issues in SQL calls
        chunk_size = 5000 
        if len(df) > chunk_size:
            chunks = np.array_split(df, math.ceil(len(df) / chunk_size))
            for i, chunk in enumerate(chunks):
                drop_into_db(table_name=table_name,
                             columns=columns,
                             values=chunk.values.tolist(),
                             append_to_prev=append_to_prev if i == 0 else True,
                             db_name=db_name)
        else:
            drop_into_db(table_name=table_name,
                         columns=columns,
                         values=df.values.tolist(),
                         append_to_prev=append_to_prev,
                         db_name=db_name)

def merge_data_files(path=None,
                    file_list=None,
                    postfix='xlsx',
                    return_df=False,
                    delete_after_merge=False,
                    drop_to_sql=False,
                    table_name=None,
                    db_name=None,
                    drop_to_excel=False,
                    excel_dest=None,
                    add_extraInfoToDf=False,
                    append_to_prev=False,
                    multithread=False):
    """
    Unified function to merge Excel or HTML files.
    """
    if file_list is None and path is not None:
        pattern = postfix if postfix.startswith('.') else f".{postfix}"
        file_list = glob.glob(os.path.join(path, f"*{pattern}"))
    
    if not file_list:
        return pd.DataFrame() if return_df else None

    dfs = []
    for f in file_list:
        try:
            if postfix.lower() in ['html', 'htm']:
                df = pd.read_html(f)[0]
            else:
                df = pd.read_excel(f)
            
            if add_extraInfoToDf:
                df['filename_info'] = get_filename(f)
                
            dfs.append(df)
        except Exception as e:
            print(f"Error reading {f}: {e}")
            # Fallback for corrupted Excel files if needed
            if postfix.lower() in ['xlsx', 'xls']:
                try:
                    save_excel(f, only_save=True)
                    df = pd.read_excel(f)
                    dfs.append(df)
                except:
                    pass

    if not dfs:
        return pd.DataFrame() if return_df else None

    final_df = pd.concat(dfs, ignore_index=True)
    final_df = final_df.astype('str')
    final_df['آخرین بروزرسانی'] = get_update_date()

    if drop_to_sql and table_name:
        _drop_data_to_db(final_df, table_name, append_to_prev=append_to_prev, 
                         db_name=db_name, multithread=multithread)

    if drop_to_excel and excel_dest:
        maybe_make_dir([os.path.dirname(excel_dest)])
        final_df.to_excel(excel_dest, index=False)

    if delete_after_merge:
        remove_excel_files(files=file_list)

    if return_df:
        return final_df

def merge_multiple_excel_sheets(path, dest, return_df=False, delete_after_merge=False, 
                                postfix='.xls', db_save=False, table=None, multithread=False):
    """Backward compatibility wrapper for merge_multiple_excel_sheets."""
    return merge_data_files(path=path, postfix=postfix, return_df=return_df, 
                           delete_after_merge=delete_after_merge, drop_to_sql=db_save, 
                           table_name=table, multithread=multithread)

def merge_multiple_html_files(path=None, return_df=False, delete_after_merge=True,
                               drop_into_sql=False, drop_to_excel=False, file_list=None,
                               add_extraInfoToDf=False):
    """Backward compatibility wrapper for merge_multiple_html_files."""
    excel_dest = os.path.join(path, 'final_df.xlsx') if drop_to_excel and path else None
    return merge_data_files(path=path, file_list=file_list, postfix='html', 
                           return_df=return_df, delete_after_merge=delete_after_merge,
                           drop_to_sql=drop_into_sql, table_name='tblArzeshAfzoodeSonati',
                           drop_to_excel=drop_to_excel, excel_dest=excel_dest,
                           add_extraInfoToDf=add_extraInfoToDf)

def merge_multiple_excel_files(path, dest, excel_name='merged', table_name='default',
                               delete_after_merge=True, return_df=False, postfix='xlsx',
                               postfix_after_merge='xlsx', drop_to_sql=False, append_to_prev=False):
    """Backward compatibility wrapper for merge_multiple_excel_files."""
    excel_dest = os.path.join(dest, excel_name, f"{table_name}.{postfix_after_merge}")
    return merge_data_files(path=path, postfix=postfix, return_df=return_df, 
                           delete_after_merge=delete_after_merge, drop_to_sql=drop_to_sql,
                           table_name=table_name, drop_to_excel=True, excel_dest=excel_dest,
                           append_to_prev=append_to_prev)

def rename_files(path, dest, prefix='.xls', postfix='.html', file_list=None, *args, **kwargs):
    if file_list is None:
        file_list = glob.glob(os.path.join(path, "*" + prefix))

    maybe_make_dir([dest])
    for i, item in enumerate(file_list):
        if 'years' in kwargs and item in kwargs['years']:
            i = kwargs['years'][item]
        
        dest_file = os.path.join(dest, f"{i}{postfix}")
        os.rename(item, dest_file)

def read_multiple_excel_files(path, postfix='xlsx', *args, **kwargs):
    """Yield files matching postfix in path and its immediate subdirectories."""
    def get_files(d):
        try:
            return [it.path for it in os.scandir(d) if it.is_file() and it.name.endswith('.' + postfix)]
        except:
            return []

    # Level 0
    for f in get_files(path): yield f
    
    # Level 1 Subdirs
    try:
        subdirs = [it.path for it in os.scandir(path) if it.is_dir()]
        for sd in subdirs:
            for f in get_files(sd): yield f
            # Level 2 Subdirs
            try:
                sub_subdirs = [it.path for it in os.scandir(sd) if it.is_dir()]
                for ssd in sub_subdirs:
                    for f in get_files(ssd): yield f
            except: pass
    except: pass

@check_if_up_to_date
def is_updated_to_save(path):
    return True

def save_row(row, tqdm_pandas, csv_name, csv_exists):
    tqdm_pandas.update(1)
    row.to_frame().transpose().to_csv(csv_name, mode='a', header=not csv_exists)
    return row

@measure_time
def read_excel_withtime(path):
    return pd.read_excel(path)

@measure_time
def save_as_csv_file(excel_file):
    with xw.App(visible=False) as app:
        wb = app.books.open(excel_file)
        sheet = wb.sheets[0] # Default to first sheet
        data = sheet.range('A1').expand().options(pd.DataFrame).value
        csv_name = f"{os.path.splitext(excel_file)[0]}.csv"
        data.to_csv(csv_name, index=False)
        wb.close()

def cleanupdf(data):
    if data is None or data.empty:
        return data
    
    # Standard cleanup replacements
    for val in [np.nan, 'nan', 'None', 'NaN', 'NaT', None]:
        data.replace(val, '0', inplace=True)
    data.fillna('0', inplace=True)

    # Identify numeric columns and convert to int-strings
    for col in data.columns:
        numeric_series = pd.to_numeric(data[col], errors='coerce')
        if numeric_series.notna().all():
            data[col] = numeric_series.replace(np.nan, 0).astype(np.int64).astype(str)
    
    data = data.astype('str')
    data.replace('0', '', inplace=True)
    data.replace('0.0', '', inplace=True)
    return data

@measure_time
def save_excel(excel_file, log=True, only_save=True, save_as_csv=False,
               save_into_sql=False, table_name=None, db_name=None, multi_process=False, append_to_prev=False):
    
    with xw.App(visible=False) as app:
        irismash = app.books.open(excel_file)
        if only_save:
            irismash.save()
            irismash.close()
            time.sleep(2) # Reduced sleep
            return

        sheet = irismash.sheets[0]
        data = sheet.range('A1').expand().options(pd.DataFrame).value
        irismash.close()

    data["آخرین بروزرسانی"] = get_update_date()
    data = cleanupdf(data)
    
    if save_as_csv:
        csv_name = f"{os.path.splitext(excel_file)[0]}.csv"
        data.to_csv(csv_name, index=False)
        
    if save_into_sql and table_name:
        _drop_data_to_db(data, table_name, append_to_prev=append_to_prev, 
                         db_name=db_name, multithread=multi_process)

def read_excel_and_drop_into_db(file, table_name, append_to_prev=False, db_name='TestDb'):
    df = read_excel_withtime(file)
    df = cleanupdf(df)
    df['آخرین بروزرسانی'] = get_update_date()
    _drop_data_to_db(df, table_name, append_to_prev=append_to_prev, db_name=db_name)

def open_and_save_excel_files(path, report_type=None, year=None,
                              merge=False, only_save=True, save_as_csv=False,
                              save_into_sql=False, table_name=None,
                              db_name=None, postfixes=['xlsx', 'xls', 'html'], append_to_prev=False,
                              multi_process=False, *args, **kwargs):
    excel_files = []
    for postfix in postfixes:
        excel_files.extend(glob.glob(os.path.join(path, f"*.{postfix}")))

    for f in excel_files:
        save_excel(f, only_save=only_save, save_as_csv=save_as_csv,
                   save_into_sql=save_into_sql, multi_process=multi_process,
                   table_name=table_name, db_name=db_name, append_to_prev=append_to_prev)

def is_updated_to_download(table_name=None, type_of='download'):
    sql_query = """SELECT * FROM tblLog WHERE table_name = '%s'
                    and update_date = (select MAX(update_date) FROM tblLog 
                    WHERE table_name = '%s')
                    and type_of = '%s'""" % (table_name, table_name, type_of)

    df = connect_to_sql(sql_query, sql_con=None, read_from_sql=True, return_df=True)

    if not df.empty and df['update_date'].iloc[0] == get_update_date():
        return True
    return False

def check_if_col_exists(df, col):
    return col in df.columns

@measure_time
def create_df(excel_files, year=None, report_type=None, type_of_report=None):
    merge_excels = []
    
    # Hardcoded logic preserved as requested (though it should ideally be refactored further)
    patterns = {
        r'C:\ezhar-temp\1392\amar_sodor_gharar_karshenasi\گزارش آمار صدور قرار کارشناسی.xlsx': None,
        r'C:\ezhar-temp\1392\imp_parvand\صد و دویست پرونده انتخابی دادستانی.xlsx': None,
        r'C:\ezhar-temp\%s\%s\Excel.xlsx' % (year, report_type): ('منبع مالیاتی', 'مالیات بر درآمد شرکت ها'),
        r'C:\ezhar-temp\%s\%s\Excel(1).xlsx' % (year, report_type): ('منبع مالیاتی', 'مالیات بر درآمد مشاغل'),
        r'C:\ezhar-temp\%s\%s\Excel(2).xlsx' % (year, report_type): ('منبع مالیاتی', 'مالیات بر ارزش افزوده'),
    }
    
    html_patterns = [
        r'C:\ezhar-temp\%s\%s\جزئیات اعتراضات و شکایات.html' % (year, report_type),
        r'C:\ezhar-temp\%s\%s\گزارش شکايات در جريان دادرسی در هيات بدوی، تجديد نظر و هم عرض.html' % (year, report_type),
        r'C:\ezhar-temp\%s\%s\جستجو.html' % (year, report_type)
    ]

    for f in excel_files:
        if f in patterns:
            df = pd.read_excel(f)
            if patterns[f]:
                col, val = patterns[f]
                if col not in df.columns:
                    df.insert(11, column=col, value=val)
            merge_excels.append(df)
        elif f in html_patterns:
            df = pd.read_html(f, flavor="html5lib")[0]
            merge_excels.append(df)

    if not merge_excels:
        return [], []

    final_df = pd.concat(merge_excels, ignore_index=True)
    final_df = cleanupdf(final_df)
    final_df['تاریخ بروزرسانی'] = get_update_date()

    return final_df.values.tolist(), final_df.columns

def highlight_excel_cells(df, rows_range, cols_range=None, saved_path=None, hex_color='4d4dff'):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    if cols_range is None:
        max_cols = ws.max_column
        cols_range = [get_column_letter(i) for i in range(1, max_cols + 1)]

    fill = styles.fills.PatternFill(patternType='solid', fgColor=styles.Color(rgb=hex_color))
    
    for i in range(1, rows_range + 2): # Header row + rows_range
        for col_letter in cols_range:
            ws[f'{col_letter}{i}'].fill = fill

    if saved_path:
        wb.save(saved_path)

def df_to_excelsheet(saved_path, dict_df, index, *args, **kwargs):
    with pd.ExcelWriter(saved_path) as writer:
        for key, dataframe in dict_df.items():
            if isinstance(key, tuple):
                sheet_name = '-'.join(str(k) for k in key)
                if 'names' in kwargs:
                    for name in kwargs['names']:
                        sheet_name = sheet_name.replace(name, '')
            else:
                sheet_name = str(key)
            
            # Excel sheet name limit is 31 chars
            dataframe.to_excel(writer, sheet_name=sheet_name[:31], index=index)

def df_to_excelworkbooks(df, g_cols, saving_path=None,
                         zipped_output={'zipped': True,
                                        'dir': r'E:\automating_reports_V2\saved_dir\arzeshafzoodeh_sonati\test\final'},
                         *args, **kwargs):
    from codeeghtesadi.utils.file_ops import zipdir 

    if saving_path is None:
        saving_path = r'E:\automating_reports_V2\saved_dir\arzeshafzoodeh_sonati\test\final\test'

    grouped_df = df.groupby(g_cols)

    for key, item in grouped_df:
        # key might be a tuple if g_cols is a list
        k_str = '-'.join(str(k) for k in key) if isinstance(key, tuple) else str(key)
        saved_dir = os.path.join(saving_path, k_str)
        maybe_make_dir([saved_dir])
        item.to_excel(os.path.join(saved_dir, f"{k_str}.xlsx"))

    if zipped_output.get('zipped'):
        maybe_make_dir([zipped_output['dir']])
        file_name = os.path.join(zipped_output['dir'], 'final.zip')
        import zipfile
        with zipfile.ZipFile(file_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipdir(saving_path, zipf)
