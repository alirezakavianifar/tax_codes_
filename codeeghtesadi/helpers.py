import tensorflow as tf
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import glob
import re
from persiantools.jdatetime import JalaliDate
from datetime import timedelta
# import tensorflow as tf
import numpy as np
import os.path
import pandas as pd
from datetime import datetime
import argparse
import jdatetime
from jdatetime import datetime, timedelta
import xlwings as xw
import shutil
import pandas as pd
import zipfile
from contextlib import contextmanager
import datetime as dt
from datetime import datetime
from functools import wraps
import functools as ft
import math
import psutil
from tqdm import tqdm
import pyodbc
import selenium
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, wait, ProcessPoolExecutor
from urllib3.exceptions import ProtocolError, MaxRetryError
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, fills
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import keyboard
# from tensorflow import keras
# from tensorflow.keras import layers
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from codeeghtesadi.constants import geck_location, set_gecko_prefs, get_remote_sql_con, get_sql_con, get_str_help, get_comm_reports, \
    get_heiat, get_lst_reports, get_all_years, get_common_years, get_str_years, get_years, get_common_reports, \
    get_comm_years, get_heiat_reports, get_server_namesV2
from codeeghtesadi.sql_queries import get_sql_mashaghelsonati, get_sql_mashaghelsonati_ghatee, get_sql_mashaghelsonati_tashkhisEblaghNoGhatee, \
    get_sql_mashaghelsonati_ghateeEblaghNashode, get_sql_mashaghelsonati_tashkhisEblaghNashode, \
    get_sql_mashaghelsonati_amadeghatee, get_sql_mashaghelsonati_amadeersalbeheiat, \
    sql_delete, create_sql_table, insert_into, get_tblupdateDate
from dateutil.parser import parse
from urllib.request import urlretrieve
from contextlib import contextmanager
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile

df = pd.DataFrame({'A': [1, 2]})

n_retries = 0
year = 0
report_type = 0

log_folder_name = 'C:\ezhar-temp'
log_excel_name = 'excel.xlsx'
log_dir = os.path.join(log_folder_name, log_excel_name)
saved_folder = geck_location(set_save_dir=False)


def retry_V1(retries=3, delay=1):

    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_exception = e
                    time.sleep(delay)
                    print(
                        f"Attempt {attempt + 1} failed: {e}. Retrying in {delay} seconds...")
            raise last_exception
        return wrapper
    return decorator


def leading_zero(x, sanim_based=False):

    # Check if code eghtesadi is sanim_based
    if sanim_based:
        if len(str(x)) == 12:
            # Pad two zeros at the beginning and return the result
            return f'00{x}'
        elif len(str(x)) == 13:
            # Pad two zeros at the beginning and return the result
            return f'0{x}'

    # Check if the length of the string representation of x is 8
    if len(str(x)) == 8:
        # Pad two zeros at the beginning and return the result
        return f'00{x}'
    # Check if the length of the string representation of x is 9
    elif len(str(x)) == 9:
        # Pad one zero at the beginning and return the result
        return f'0{x}'
    # Check if the length of the string representation of x is 12
    elif len(str(x)) == 12:
        # Return the first 10 characters of the string
        return str(x[:10])
    else:
        # If none of the above conditions are met, return 0 (invalid input)
        return str(x)


def measure_time(func):
    # Use the @wraps decorator to preserve the original function's metadata
    @wraps(func)
    def wrapper(*args, **kwargs):
        # Record the starting time using time.process_time()
        time1 = time.process_time()

        # Print a message indicating the start of the function and the current time
        print(
            f'***********************************************************************\n{func.__name__} started at {datetime.now().strftime("%H:%M:%S")}\n')

        # Call the original function and store its result
        res = func(*args, **kwargs)

        # Record the ending time using time.process_time()
        time2 = time.process_time()

        # Calculate the total time taken for the function execution
        all_time = time2 - time1

        # Print a message indicating the end of the function, the current time, and the time taken
        print(
            f'***********************************************************************\n {func.__name__} finished at {datetime.now().strftime("%H:%M:%S")}, time taken is {all_time}')

        # Return the result of the original function
        return res

    # Return the wrapper function
    return wrapper


@measure_time
def wrap_func(func, *args, **kwargs):
    res = func(*args, **kwargs)
    return res


def wrap_a_wrapper(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            result, info = func(*args, **kwargs)

            if ((not info['success']) and (not info['keep_alive'])):
                raise Exception
            elif ((not info['success']) and (info['keep_alive'])):
                return driver, info
            else:
                return result, info
            # else:
            #     raise Exception
        except:
            raise Exception

    return wrapper


# @wrap_a_wrapper
def wrap_it_with_params(num_tries=3, time_out=20,
                        driver_based=False, detrimental=True,
                        clean_up=False, keep_alive=False, record_process_details=False):

    def wrap_it(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            nonlocal num_tries
            result = None
            i = 0
            while (True):
                if i < num_tries:
                    # Update success information in kwargs
                    if 'info' in kwargs:
                        kwargs['info']['success'] = True
                    else:
                        kwargs['info'] = {}
                        kwargs['info']['success'] = True

                    # Log function call
                    print(
                        f'function {func.__name__} is called at {datetime.now()}')
                    try:
                        # Measure function execution time
                        start = time.process_time()
                        if record_process_details:
                            df = pd.DataFrame(
                                {'pid': [os.getpid()], 'name': [func.__name__]})
                            drop_into_db(table_name='tblAliveProcesses',
                                         columns=df.columns.tolist(),
                                         values=df.values.tolist(),
                                         db_name='test',
                                         append_to_prev=True)

                        result, info = func(*args, **kwargs)
                        end = time.process_time()
                        elapsed_time = end - start

                        # Check if elapsed time exceeds the timeout
                        if elapsed_time > time_out:
                            result = None
                        if 'success' not in info:
                            info['success'] = True
                        # Check for unsuccessful execution or termination conditions
                        if ((result is None) or ((not info['success']) and (not keep_alive))):
                            raise Exception
                        else:
                            # Log function completion
                            print(
                                f'function {func.__name__} is finished at {datetime.now()}')
                            break
                    except Exception as e:
                        if isinstance(e, (ProtocolError, MaxRetryError, TypeError)):
                            num_tries = 1000
                            return
                        # Handle exceptions and retry
                        print(
                            f'Exception occurred in function {func.__name__}, the error is {e}')
                        time.sleep(1)
                        i += 1
                        if i < num_tries:
                            print(
                                f'retrying function {func.__name__} for {i} times at {datetime.now()}')
                else:
                    try:
                        # Update info for unsuccessful execution
                        kwargs['info']['keep_alive'] = keep_alive
                        kwargs['info']['success'] = False

                        # Handle driver-based scenarios
                        if driver_based:
                            if 'driver' in kwargs:
                                driver = kwargs['driver']
                            else:
                                for item in args:
                                    if isinstance(item, selenium.webdriver.firefox.webdriver.WebDriver):
                                        driver = item
                            result = driver

                        # Perform clean-up actions
                        if clean_up:
                            driver, info = cleanup(
                                driver=driver, info=kwargs['info'])

                        # Perform detrimental actions
                        if detrimental:
                            driver.quit()

                        # Terminate function if keep_alive is False
                        if not keep_alive:
                            # raise Exception
                            pass

                        return driver, kwargs['info']
                    except Exception as e:
                        # Handle exceptions during cleanup
                        raise Exception

            return result, info

        return wrapper
    return wrap_it


def add_one_day(current_date):
    current_date = jdatetime.datetime.strptime(current_date, "%Y%m%d")
    current_date = current_date + jdatetime.timedelta(days=1)
    return current_date.strftime("%Y%m%d")


def wrap_it_with_paramsv1(num_tries=3, time_out=10,
                          driver_based=False, detrimental=True,
                          clean_up=False, keep_alive=False):

    def wrap_it(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            nonlocal num_tries
            result = None
            i = 0

            # Set default info dictionary if not provided
            if not 'info' in kwargs:
                kwargs['info'] = {}
                kwargs['info']['success'] = True

            # Retry loop
            while True:
                if i < num_tries:
                    print(
                        f'function {func.__name__} is called at {datetime.now()}')
                    try:
                        start = time.process_time()
                        kwargs['info']['success'] = True
                        result, info = func(*args, **kwargs)
                        end = time.process_time()
                        print(
                            f'function {func.__name__} is finished at {datetime.now()}')
                        break
                    except Exception as e:
                        if isinstance(e, (ProtocolError, MaxRetryError, TypeError)):
                            num_tries = 1000
                            return
                        print(
                            f'Exception occurred in function {func.__name__}, the error is : {e}')
                        time.sleep(2)
                        i += 1
                        if i < num_tries:
                            print(
                                f'retrying function {func.__name__} for {i} times at {datetime.now()}')
                else:
                    try:
                        kwargs['info']['success'] = False
                        if driver_based:
                            if 'driver' in kwargs:
                                driver = kwargs['driver']
                            else:
                                for item in args:
                                    if isinstance(item, selenium.webdriver.firefox.webdriver.WebDriver):
                                        kwargs['driver'] = item
                                        driver = item
                        if clean_up:
                            driver, info = cleanup(
                                driver=driver, info=kwargs['info'])
                        if detrimental:
                            driver.close()
                        return kwargs['driver'], kwargs['info']
                    except Exception as e:
                        print(e)
                        return kwargs['driver'], kwargs['info']

            return result, info

        return wrapper

    return wrap_it


@wrap_it_with_params(15, 10, True, False, True, False)
def find_obj_login(driver, info):

    # Use WebDriverWait to wait for up to 3 seconds for the specified element to be clickable.
    WebDriverWait(driver, 3).until(
        EC.element_to_be_clickable(
            (By.ID, 'OBJ')
        )
    )
    # Return the driver and info parameters.
    return driver, info


# Wrapper decorator that wraps another function with specific parameters.
# @wrap_a_wrapper
@wrap_it_with_params(1, 50, True, False, False, False)
def cleanup(driver, info, close_driver=False):

    if isinstance(driver, tuple):
        driver = driver[0]
        driver.switch_to.default_content()
    else:
        driver.switch_to.default_content()

    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located(
            (By.ID, 'logoutLink')))
    driver.find_element(
        By.ID,
        'logoutLink'
    ).click()

    if close_driver:
        driver.close()

    return driver, info


def check_driver_health(driver):

    if driver is None:
        raise Exception
    else:
        return driver


# Decorator that retries a function up to 5 times in case of an exception.
def retry(func):

    def try_it(*args, **kwargs):
        global n_retries
        try:
            result = func()
            return result
        except Exception as e:
            n_retries += 1
            print(e)
            if n_retries < 5:
                try_it()

    return try_it


# Decorator that retries a function up to 5 times, closing the driver on exception.
def retryV1(func):

    def try_it(*args, **kwargs):
        global n_retries
        try:
            result = func(*args, **kwargs)
            if isinstance(result, tuple):
                raise Exception
            return result
        except Exception as e:
            n_retries += 1
            result[0].close()
            if n_retries < 5:
                try_it(*args, **kwargs)

    return try_it


# Decorator that retries a function up to 5 times, closing the driver on exception.
def retry_with_arguments(driver=None):

    def retry(func):
        def try_it(*args, **kwargs):
            global n_retries
            try:
                result = func(*args, **kwargs)
                return result
            except Exception as e:
                print('error occurred. please pay close attention')
                n_retries += 1
                if n_retries < 5:
                    driver.close()
                    time.sleep(4)
                    try_it()

        return try_it

    return retry


def maybe_make_dir(directory):

    # Check if a single directory path is provided
    if isinstance(directory, str):
        directory = [directory]  # Convert to a list for uniform processing

    # Iterate through each directory path
    for item in directory:
        # Check if the directory already exists
        if not os.path.exists(item):
            # If not, create the directory and any necessary parent directories
            os.makedirs(item)


@measure_time
def read_multiple_excel_sheets(path, sheet_name=None):

    # Read all sheets from the Excel file into a dictionary of DataFrames
    df_dict = pd.read_excel(path, sheet_name=None)

    # Initialize an empty list to store individual DataFrames
    df_list = []

    # Iterate through each sheet in the dictionary
    for key in df_dict:
        # Append each DataFrame to the list
        df_list.append(df_dict[key])

    # Concatenate all DataFrames in the list into a single DataFrame
    concatenated_df = pd.concat(df_list)

    # Return the final concatenated DataFrame
    return concatenated_df

# Decorator function for logging function start and completion


def log_the_func(log_file_path, *log_args, **log_kwargs):
    def wrapper(func):
        @wraps(func)
        def try_it(*args, **kwargs):
            # Log function start with optional custom message
            if ('field' in kwargs and kwargs['field'] is not None):
                if 'soratmoamelat' in log_kwargs:
                    kwargs['field']('Starting the function %s for "%s" \n' % (
                        func.__name__, kwargs['selected_option_text']))
                else:
                    kwargs['field']('Starting the function %s" \n' % (
                        func.__name__))

            # Check if log_file_path is provided for file logging
            if log_file_path is not None:
                # Log function start to the specified log file
                with open(log_file_path, 'w') as f:
                    f.write('Starting the function %s' % func.__name__)
                    f.write('\n')
                print('Starting the function %s' % func.__name__)

                # Execute the wrapped function and capture the result
                result = func(*args, **kwargs)

                # Check if the result is a tuple and return it
                if isinstance(result, tuple):
                    try:
                        raise Exception
                    except:
                        return result

                # Log function completion with optional custom message
                if 'field' in kwargs and kwargs['field'] is not None:
                    kwargs['field'](
                        'The function %s completed successfully \n' % func.__name__)
                print('The function %s completed successfully' % func.__name__)

                # Log function completion to the specified log file
                with open(log_file_path, 'w') as f:
                    f.write('The function %s completed successfully' %
                            func.__name__)
                    f.write('\n')

            return result
        return try_it
    return wrapper

# Example function decorated with log_the_func


@log_the_func('none')
def remove_excel_files(files=[],      # List to store file paths to be removed
                       pathsave=None,  # Unused parameter, can be removed if not needed
                       file_path=None,  # Path where files are located
                       # File types to be removed (default is 'html')
                       postfix=['html'],
                       *args,  # Additional positional arguments (unused)
                       **kwargs):  # Additional keyword arguments (unused)

    # If a specific file_path is provided, find files with specified postfix in that path
    if file_path is not None:
        for item in postfix:
            files.extend(glob.glob(file_path + "/*" + item))

    # Iterate through the list of files to be removed
    for f in files:
        # Check if the file exists
        if os.path.exists(f):
            # If the file exists, remove it
            os.remove(f)


def merge_multiple_excel_sheets(path,
                                dest,
                                return_df=False,
                                delete_after_merge=False,
                                postfix='.xls',
                                db_save=False,
                                table=None,
                                multithread=False):

    # List to store DataFrames from individual sheets
    lst = []

    # Get a list of Excel files in the specified path with the given file extension
    file_list = glob.glob(path + "/*" + postfix)

    # Iterate through each Excel file
    for item in file_list:
        # Open Excel file
        xl = pd.ExcelFile(item)

        # Read the first sheet into a DataFrame
        df = pd.read_excel(item, sheet_name=xl.sheet_names[0])
        df = df.astype('str')

        # Save the DataFrame to the database (if specified)
        drop_into_db(table_name=table,
                     columns=df.columns.tolist(),
                     values=df.values.tolist(),
                     append_to_prev=False)

        # Function to read each sheet (starting from the second) and save to the database
        def df_to_db(sheet):
            df = pd.read_excel(item, sheet_name=sheet)
            df = df.astype('str')
            drop_into_db(table_name=table,
                         columns=df.columns.tolist(),
                         values=df.values.tolist(),
                         append_to_prev=True,
                         del_tbl='no',
                         create_tbl='no')

        # Process sheets sequentially or using multithreading
        if not multithread:
            # Iterate through sheets and save to the database
            for sheet in xl.sheet_names[1:]:
                df_to_db(sheet)
        else:
            # Use ThreadPoolExecutor for multithreading
            executor = ThreadPoolExecutor(res - 1)
            jobs = [executor.submit(df_to_db, sheet)
                    for sheet in xl.sheet_names[1:]]
            wait(jobs)

    # Optionally return the merged DataFrame
    if return_df:
        # Combine DataFrames from all sheets into a single DataFrame
        merged_df = pd.concat(lst, ignore_index=True)
        return merged_df

    # Optionally delete original Excel files after merging
    if delete_after_merge:
        for item in file_list:
            os.remove(item)


def rename_files(path,
                 dest,
                 prefix='.xls',
                 postfix='.html',
                 file_list=None,
                 *args,
                 **kwargs):
    # If file_list is not provided, use glob to get all files matching the prefix in the source directory
    if file_list is None:
        file_list = glob.glob(path + "/*" + prefix)

    # Iterate through the list of files and rename them
    for i, item in enumerate(file_list):
        # Check if 'years' is specified in keyword arguments
        if 'years' in kwargs:
            # If the current file is in the 'years' dictionary, update the index
            if item in kwargs['years']:
                i = kwargs['years'][item]

        # Create the destination path for the renamed file
        dest1 = os.path.join(dest, '%s%s' % (i, postfix))

        # Rename the file
        os.rename(item, dest1)


def get_filename(item, patterns='[\w-]+?(?=\.)'):

    # Use the re.search function to find the first match of the pattern in the item
    a = re.search(patterns, item)

    # Return the matched portion of the string (file name)
    return a.group()


def merge_multiple_html_files(path=None,
                              return_df=False,
                              delete_after_merge=True,
                              drop_into_sql=False,
                              drop_to_excel=False,
                              file_list=None,
                              add_extraInfoToDf=False):

    try:
        # If file_list is not provided, fetch all HTML files in the specified path
        if file_list is None:
            file_list = glob.glob(os.path.join(path, "*.html"))

        merge_excels = []
        # Iterate through each HTML file and read it into a DataFrame
        for item in file_list:
            df = pd.read_html(item)[0]
            # If add_extraInfoToDf is True, add extra information to the DataFrame
            if add_extraInfoToDf:
                filename = get_filename(item)
                df['سال عملکرد'] = filename
            merge_excels.append(df)

        # Concatenate all DataFrames into a final DataFrame
        final_df = pd.concat(merge_excels)
        final_df = final_df.astype('str')

        # Add the update date to the DataFrame
        final_df['آخرین بروزرسانی'] = get_update_date()

        # If drop_into_sql is True, drop the DataFrame into a SQL table
        if drop_into_sql:
            drop_into_db('tblArzeshAfzoodeSonati',
                         final_df.columns.tolist(),
                         final_df.values.tolist(),
                         append_to_prev=False)

        # If drop_to_excel is True, save the DataFrame to an Excel file
        if drop_to_excel:
            excel_path = os.path.join(path, 'final_df.xlsx')
            final_df.to_excel(excel_path)

        # If delete_after_merge is True, remove the source HTML files
        if delete_after_merge:
            remove_excel_files(files=file_list)

        # If return_df is True, return the merged DataFrame; otherwise, delete it
        if return_df:
            return final_df
        else:
            del final_df

    except Exception as e:
        # Print any exception that occurs during the process
        print(e)
        return


def read_multiple_excel_files(path, postfix='xls', *args, **kwargs):

    final_file = []
    dirs = []
    dirs.extend([it.path for it in os.scandir(path) if it.is_dir()])
    final_dirs = dirs
    for item in dirs:
        final_dirs.extend([it.path for it in os.scandir(item) if it.is_dir()])
    for item in final_dirs:
        files = list_files(item, postfix)
        for file in files:
            yield file


def merge_multiple_excel_files(path,
                               dest,
                               excel_name='merged',
                               table_name='default',
                               delete_after_merge=True,
                               return_df=False,
                               postfix='xlsx',
                               postfix_after_merge='xlsx',
                               drop_to_sql=False,
                               append_to_prev=False):
    # Construct the destination path
    dest = os.path.join(dest, excel_name)
    maybe_make_dir([dest])
    dest = os.path.join(dest, table_name + '.' + postfix_after_merge)

    # Get the list of excel files in the specified path
    excel_files = glob.glob(os.path.join(path, "*.%s" % postfix))
    merge_excels = []

    # Read each excel file into a DataFrame and append to the list
    for f in excel_files:
        try:
            df = pd.read_excel(f)
            merge_excels.append(df)
        except Exception as e:
            save_excel(f, log=False)
            df = pd.read_excel(f)
            merge_excels.append(df)

    # Merge all DataFrames into one
    final_df_all_fine_grained = pd.concat(merge_excels)
    final_df_all_fine_grained = final_df_all_fine_grained.astype('str')
    final_df_all_fine_grained['تاریخ بروزرسانی'] = get_update_date()

    # Optionally, delete original excel files
    if delete_after_merge:
        remove_excel_files(files=excel_files)

    # Optionally, drop merged data into a SQL database
    if drop_to_sql:
        drop_into_db(table_name=table_name,
                     columns=final_df_all_fine_grained.columns.tolist(),
                     values=final_df_all_fine_grained.values.tolist(),
                     append_to_prev=True)

    # Save the merged DataFrame to an excel file
    final_df_all_fine_grained.to_excel(dest, index=False)

    print('Merging was done successfully')

    # Optionally, return the merged DataFrame
    if return_df:
        return final_df_all_fine_grained
    else:
        # Delete the merged DataFrame if not returned
        del final_df_all_fine_grained


@retry
def input_info():
    # Initialize the argument parser
    parser = argparse.ArgumentParser()

    # Define command-line arguments with their respective help messages
    parser.add_argument('--reportTypes', type=str, help=get_str_help())
    parser.add_argument('--years', type=str, help=get_str_years())
    parser.add_argument('--s',
                        type=str,
                        nargs='?',
                        default='not-s',
                        help='types:\nt = True\nf = False\n')
    parser.add_argument('--d',
                        type=str,
                        nargs='?',
                        default='not-d',
                        help='types:\nt = True\nf = False\n')
    parser.add_argument('--c',
                        type=str,
                        nargs='?',
                        default='not-c',
                        help='types:\nt = True\nf = False\n')

    # Parse the command-line arguments
    args = parser.parse_args()

    # Extract individual arguments from the parsed results
    selected_report_types = args.reportTypes.split(',')  # ['1','2','3','4']
    selected_years = args.years.split(',')

    # Get various data needed for processing
    all_years = get_all_years()
    comm_reports = get_comm_reports()
    comm_years = get_comm_years()
    heiat = get_heiat()

    # Extract arguments related to actions
    dump_to_sql = args.d
    create_reports = args.c
    scrape = args.s

    # Initialize lists to store selected report types and years
    new_reportTypes = []
    new_years = []

    # Retrieve various data lists
    report_types = get_lst_reports()
    years = get_years()
    common_reports = get_common_reports()
    common_years = get_common_years()
    heiat_reports = get_heiat_reports()

    # Process selected report types
    if selected_report_types.count(comm_reports):
        for report_type in common_reports:
            new_reportTypes.append(report_type)

    elif selected_report_types.count(heiat):
        for report_type in heiat_reports:
            new_reportTypes.append(report_type[0])

    else:
        for report_type in report_types:
            if selected_report_types.count(report_type[1]):
                new_reportTypes.append(report_type[0])

    # Process selected years
    if selected_years.count(all_years):
        for year in years:
            new_years.append(year[0])

    elif selected_years.count(comm_years):
        for year in common_years:
            new_years.append(year[0])

    else:
        for year in years:
            if selected_years.count(year[1]):
                new_years.append(year[0])

    # Return the processed data
    return new_reportTypes, new_years, scrape, dump_to_sql, create_reports


def get_update_date(date=None, delimiter=None):
    # If no date is provided, use today's date in Jalali calendar
    if date is None:
        x = jdatetime.date.today()

        # Format month with leading zero if it's a single-digit month
        if len(str(x.month)) == 1:
            month = '0%s' % str(x.month)
        else:
            month = str(x.month)

        # Format day with leading zero if it's a single-digit day
        if len(str(x.day)) == 1:
            day = '0%s' % str(x.day)
        else:
            day = str(x.day)

        year = str(x.year)

    # If a Gregorian date is provided, convert it to Jalali calendar
    else:
        gdate = jdatetime.GregorianToJalali(int(date[0]), int(date[1]),
                                            int(date[2]))
        year = str(gdate.jyear)
        month = str(gdate.jmonth)
        day = str(gdate.jday)

        # Format month with leading zero if it's a single-digit month
        if len(str(month)) == 1:
            month = '0%s' % str(month)
        else:
            month = str(month)

        # Format day with leading zero if it's a single-digit day
        if len(str(day)) == 1:
            day = '0%s' % str(day)
        else:
            day = str(day)

    # Check if a delimiter is specified and format the date accordingly
    if delimiter is not None:
        update_date = year + delimiter + month + delimiter + day
    else:
        update_date = year + month + day

    return update_date


def check_update(func):
    # Decorator to check if a database table has been updated before executing a function

    @wraps(func)
    def inner_wrapper(*args, **kwargs):
        # Function to check and handle updates before executing the decorated function

        # Get the last update date from the table
        sql_query = get_tblupdateDate(kwargs['table_name'])
        date = connect_to_sql(
            sql_query, sql_con=get_sql_con(database='testdbV2'), read_from_sql=True, return_df=True)

        # Compare the last update date with the current update date
        if date.iloc[0][0] == get_update_date():
            # If the table has already been updated, print a message and return None
            result = None
            print('Table already updated. Skipping function execution.')
        else:
            # If the table has not been updated, proceed with executing the decorated function
            result = func(*args, **kwargs)

        return result

    return inner_wrapper


def record_keys(max_num=4, until='esc'):
    # Define a list of accepted letters and digits
    accepted_letters = ['a', 'A', 'b', 'B', 'c', 'C', 'd', 'D', 'e', 'E', 'f', 'F', 'g', 'G', 'h', 'H', 'i', 'I', 'j', 'J', 'k', 'K', 'l', 'L', 'm', 'M',
                        'n', 'N', 'o', 'O', 'p', 'P', 'q', 'Q', 'r', 'R', 's', 'S', 't', 'T', 'u', 'U', 'v', 'V',
                        'w', 'W', 'x', 'X', 'y', 'Y', 'z', 'Z', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9']

    # Initialize variables
    accepted = False
    recorded = keyboard.record(until=until)
    lst = []

    # Iterate over the recorded keys and filter out accepted ones
    for index, item in enumerate(recorded):
        if index % 2 == 0:
            if item.name in accepted_letters:
                lst.append(item.name)

    # Check if the number of recorded keys is equal to max_num
    if len(lst) != max_num:
        return accepted, 'null'

    # Concatenate the recorded keys into a string
    str_result = ''
    accepted = True
    for item in lst:
        str_result += item

    return accepted, str_result


def time_it(log=False, tbl_name='default', num_runs=12,
            db={'db_name': 'testdbV2',
                'tbl_name': 'tblLog',
                'append_to_prev': False}):

    def wrapper(func):
        @wraps(func)
        def inner_func(*args, **kwargs):
            nonlocal tbl_name
            nonlocal log
            func_name = func.__name__
            nonlocal num_runs

            # Check if num_runs is specified in kwargs, if so, override the default value.
            if 'num_runs' in kwargs:
                num_runs = kwargs['num_runs']

            if log:
                # If logging is enabled, check for additional kwargs to customize the logging.
                if 'table_name' in kwargs:
                    tbl_name = kwargs['table_name']
                else:
                    tbl_name = args[2]

                if 'type_of' in kwargs:
                    type_of = kwargs['type_of']
                else:
                    type_of = 'nan'

            start = time.time()

            def tryit():
                nonlocal num_runs
                try:
                    # Attempt to execute the decorated function.
                    result = func(*args, **kwargs)

                    # Check if the result is a tuple, and if so, raise an exception.
                    if isinstance(result, tuple):
                        kwargs['index'] = result[0]
                        raise Exception

                    return result

                except Exception as e:
                    print(e)
                    num_runs -= 1
                    x = num_runs > 0

                    # If there are more runs remaining, recursively retry the function.
                    if x:
                        tryit(*args, **kwargs)
                    else:
                        raise Exception

            # Execute the decorated function and capture the results.
            final_results = tryit()

            # If 'only_schema' is specified and True, do not log the operation.
            if 'only_schema' in kwargs:
                if kwargs['only_schema']:
                    log = False

            # Log the operation into the specified database table.
            if log:
                df_all = pd.DataFrame({
                    'table_name': [tbl_name],
                    'update_date': [get_update_date()],
                    'type_of': [type_of]
                })

                drop_into_db(table_name=db['tbl_name'],
                             columns=df_all.columns.tolist(),
                             values=df_all.values.tolist(),
                             append_to_prev=db['append_to_prev'],
                             db_name=db['db_name'])

            end = time.time()

            try:
                # Print information about the completed function.
                if 'connect_type' in kwargs:
                    print(kwargs['connect_type'])
                print('The function %s completed successfully' % func_name)
                time_taken = str('%.2f' % (end-start))
                print(func.__name__ + ' took ' + time_taken +
                      ' seconds')

                # Return the final results, if any.
                if final_results is not None:
                    return final_results
                return

            except Exception as e:
                print(e)
                return

        return inner_func

    return wrapper


def make_dir_if_not_exists(paths):
    for path in paths:
        # Check whether the specified path exists or not
        isExist = os.path.exists(path)

        if not isExist:
            # Create a new directory because it does not exist
            os.makedirs(path)


def check_if_up_to_date(func):

    @wraps(func)
    def try_it(*args, **kwargs):
        if 'log' in kwargs:
            if kwargs['log'] == False:
                result = func(*args, **kwargs)
                return result
        current_date = int(get_update_date())
        func_name = func.__name__
        if func_name == 'is_updated_to_save':
            type_of = 'save_excel'
        elif func_name == 'is_updated_to_download':
            type_of = 'download_excel'
        else:
            type_of = 'save_excel'

        df = connect_to_sql

        df = pd.read_excel(log_dir)
        check_date = df['date'].where((df['file_name'] == args[0])
                                      & (df['type_of'] == type_of)).max()

        if not math.isnan(check_date):

            if (int(check_date) == int(current_date)
                    and func_name != 'save_excel'):
                print('The %s have already been logged' % args[0])
                result = func(*args, **kwargs)
                return result

            elif (int(check_date) != int(current_date)
                  and func_name == 'save_excel'):
                print('opening excel')
                result = func(*args, **kwargs)
                return result
            else:
                return False

        elif (math.isnan(check_date) and func_name == 'is_updated_to_save'):
            return False

        elif (math.isnan(check_date)
              and func_name == 'is_updated_to_download'):
            return False

        else:

            result = func(*args, **kwargs)
            return result

            # return print('The file %s has already been updated \n' % args[0])

    return try_it


def log_it(func):

    @wraps(func)
    def try_it(*args, **kwargs):
        if 'log' in kwargs:
            if kwargs['log'] == False:
                result = func(*args, **kwargs)
                return result
        print('log_it initialized')
        d1 = datetime.now()
        type_of = func.__name__

        if (func.__name__ == 'save_excel'):

            print('opening %s for saving' % args[0])

        result = func(*args, **kwargs)

        c_date = get_update_date()

        if type_of == 'download_excel':

            df_1 = pd.DataFrame([[result, c_date, type_of]],
                                columns=['file_name', 'date', 'type_of'])

        else:

            df_1 = pd.DataFrame([[args[0], c_date, type_of]],
                                columns=['file_name', 'date', 'type_of'])

        # create excel file for logging if it does not already exist
        if not os.path.exists(log_dir):

            df_1.to_excel(log_dir)

        else:

            df_2 = pd.read_excel(log_dir, index_col=0)

            df_3 = pd.concat([df_1, df_2])

            remove_excel_files([log_dir])

            df_3.to_excel(log_dir)

        d2 = datetime.now()
        d3 = (d2 - d1).total_seconds() / 60

        if type_of == 'download_excel':

            print('it took %s minutes for the %s to be logged' %
                  ("%.2f" % d3, kwargs['type_of_excel']))

        else:

            print('it took %s minutes for the %s to be logged' %
                  ("%.2f" % d3, args[0]))

        print(
            '***********************************************************************\n'
        )

        return result

    return try_it


@check_if_up_to_date
def is_updated_to_save(path):
    return True

# Define a function to save each row


def save_row(row, tqdm_pandas, csv_name, csv_exists):
    tqdm_pandas.update(1)
    # Append the row to the CSV file
    row.to_frame().transpose().to_csv(csv_name, mode='a', header=not csv_exists)
    return row


@measure_time
def read_excel_withtime(path):
    df = pd.read_excel(path)
    return df


@measure_time
def save_as_csv_file(excel_file):
    irismash = xw.Book(excel_file)
    # Replace 'Sheet1' with the actual sheet name
    sheet = irismash.sheets['Sheet1']
    # Read data from the sheet into a pandas DataFrame
    data = sheet.range('A1').expand().options(pd.DataFrame).value
    # Save the DataFrame to a CSV file
    csv_name = f"{excel_file.split('.')[0]}.csv"
    data.to_csv(csv_name, index=False)
    irismash.app.quit()


def split_list(lst, chunk_size):
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

# @check_if_up_to_date
# @log_it


def cleanupdf(data):
    data.fillna(0, inplace=True)
    data.replace('nan', '0', inplace=True)
    data.replace('None', '0', inplace=True)
    data.replace('NaN', '0', inplace=True)
    data.replace('NaT', '0', inplace=True)
    # Identify numeric columns
    numeric_columns = data.apply(lambda x: pd.to_numeric(
        x, errors='coerce').notna().all())
    numeric_columns = data.columns[numeric_columns].tolist()
    data[numeric_columns] = data[numeric_columns].apply(
        np.float64).apply(np.int64)
    data = data.astype('str')
    data.replace('0', '', inplace=True)

    return data


@measure_time
def save_excel(excel_file, log=True, only_save=True, save_as_csv=False,
               save_into_sql=False, table_name=None, db_name=None, multi_process=False, append_to_prev=False):
    # irismash = xw.Book(excel_file, visible=False)
    # Open the Excel application
    app = xw.App(visible=False)
    # Open the Excel file
    irismash = app.books.open(excel_file)
    if only_save:
        irismash.save()
        irismash.app.quit()
        time.sleep(8)
        return
    # Save the DataFrame to a CSV file
    csv_name = f"{excel_file.split('.')[0]}.csv"
    # Replace 'Sheet1' with the actual sheet name
    sheet = irismash.sheets['Sheet1']
    # Read data from the sheet into a pandas DataFrame
    data = sheet.range('A1').expand().options(pd.DataFrame).value
    irismash.app.quit()
    data["آخرین بروزرسانی"] = get_update_date()
    data = cleanupdf(data)
    if save_as_csv:
        data.to_csv(csv_name, index=False)
    if save_into_sql:
        if multi_process:
            data = np.array_split(data, 40)
            drop_into_db(table_name=table_name,
                         columns=data[0].columns.tolist(),
                         values=data[0].values.tolist(),
                         append_to_prev=append_to_prev,
                         db_name=db_name)
            append_to_prev = True
            with ThreadPoolExecutor(39) as executor:
                jobs = [executor.submit(drop_into_db, table_name,
                                        item.columns.tolist(),
                                        item.values.tolist(),
                                        append_to_prev,
                                        db_name
                                        )
                        for index, item in tqdm(enumerate(data[1:]))]
                wait(jobs)

        else:
            data = np.array_split(data, 10000)
            append_to_prev = False
            for item in tqdm(data):
                drop_into_db(table_name=table_name,
                             columns=item.columns.tolist(),
                             values=item.values.tolist(),
                             append_to_prev=append_to_prev,
                             db_name=db_name)
                append_to_prev = True


def read_excel_and_drop_into_db(file, table_name,
                                append_to_prev=False, db_name='TestDb'):

    df = read_excel_withtime(file)
    df = cleanupdf(df)
    df['آخرین بروزرسانی'] = get_update_date()
    drop_into_db(table_name=table_name,
                 columns=df.columns.tolist(),
                 values=df.values.tolist(),
                 append_to_prev=append_to_prev,
                 db_name='TestDb')
    append_to_prev = True


def open_and_save_excel_files(path, report_type=None, year=None,
                              merge=False, only_save=True, save_as_csv=False,
                              save_into_sql=False, table_name=None,
                              db_name=None, postfixes=['xlsx', 'xls', 'html'], append_to_prev=False,
                              multi_process=False, *args, **kwargs):
    # Get a list of all Excel files in the specified path
    excel_files = []

    for postfix in postfixes:
        excel_files.extend(glob.glob(os.path.join(path, f"*.{postfix}")))

    # Iterate through each Excel file in the list
    for f in excel_files:
        # Call the save_excel function to save the Excel file
        save_excel(f, only_save=only_save, save_as_csv=save_as_csv,
                   save_into_sql=save_into_sql, multi_process=multi_process,
                   table_name=table_name, db_name=db_name, append_to_prev=append_to_prev)


def is_updated_to_download(table_name=None, type_of='download'):

    # Construct SQL query to retrieve the latest update for the specified table and type
    sql_query = """SELECT * FROM tblLog WHERE table_name = '%s'
                    and update_date = (select MAX(update_date) FROM tblLog 
                    WHERE table_name = '%s')
                    and type_of = '%s'""" % (table_name, table_name, type_of)

    # Connect to the database and execute the SQL query
    df = connect_to_sql(
        sql_query, sql_con=get_sql_con(database='testdbV2'), read_from_sql=True, return_df=True)

    # Check if there are rows returned by the query
    if int(df.count().iloc[0]) > 0:
        # Check if the update date of the first row matches the current update date
        if df['update_date'].loc[0] == get_update_date():
            # The table has been updated and is ready for download
            return True

    # The table has not been updated or does not meet the criteria
    return False


@check_if_up_to_date
def is_updated_to_save(path):

    return True


def check_if_col_exists(df, col):

    if col in df:
        return True
    else:
        return False


FIREFOX_PROFILE_PATH = (
    r"C:\Users\alkav\AppData\Roaming\Mozilla\Firefox\Profiles\q0uj5wit.taxgov"
)


@contextmanager
def init_driver(
    pathsave,
    driver_type='firefox',
    headless=False,
    prefs={'maximize': False, 'zoom': '1.0'},
    driver=None,
    info=None,
    use_proxy=False,
    disable_popups=False,
    *args,
    **kwargs
):

    if info is None:
        info = {}

    driver = None

    if driver_type == 'chrome':
        # ---------- Chrome (unchanged) ----------
        options = Options()
        options.add_argument("start-maximized")
        options.add_experimental_option(
            "prefs", {"download.default_directory": pathsave}
        )

        driver = webdriver.Chrome(
            executable_path=geck_location(driver_type='chrome'),
            options=options
        )

    else:
        # ---------- Firefox ----------
        if headless:
            raise RuntimeError(
                "Headless Firefox cannot be used with persistent profiles"
            )

        if not os.path.isdir(FIREFOX_PROFILE_PATH):
            raise RuntimeError("Firefox profile path does not exist")

        profile = FirefoxProfile(FIREFOX_PROFILE_PATH)
        options = FirefoxOptions()

        # ---------- Download preferences ----------
        profile.set_preference('browser.download.folderList', 2)
        profile.set_preference('browser.download.dir', pathsave)
        profile.set_preference(
            'browser.download.manager.showWhenStarting', False)

        profile.set_preference(
            'browser.helperApps.neverAsk.saveToDisk',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream'
        )
        profile.set_preference('browser.helperApps.alwaysAsk.force', False)

        # ---------- UI / Zoom ----------
        profile.set_preference(
            "layout.css.devPixelsPerPx",
            prefs.get('zoom', '1.0')
        )

        # ---------- Proxy ----------
        if use_proxy:
            proxy_address = "10.52.3.128"
            profile.set_preference("network.proxy.type", 1)
            profile.set_preference("network.proxy.http", proxy_address)
            profile.set_preference("network.proxy.http_port", 8080)
            profile.set_preference("network.proxy.ssl", proxy_address)
            profile.set_preference("network.proxy.ssl_port", 8080)

        # ---------- Popups ----------
        if disable_popups:
            profile.set_preference("dom.popup_maximum", 0)
            profile.set_preference("dom.popup_allowed_events", "")

        driver = webdriver.Firefox(
            service=FirefoxService(geck_location()),
            options=options,
            firefox_profile=profile
        )

        if prefs.get('maximize'):
            driver.maximize_window()

    try:
        yield driver

    finally:
        if driver:
            driver.quit()


@measure_time
def create_df(excel_files, year=None, report_type=None, type_of_report=None):

    merge_excels = []
    # Convert excel files into dataframes
    for f in excel_files:
        if f == 'C:\\ezhar-temp\\1392\\amar_sodor_gharar_karshenasi\\گزارش آمار صدور قرار کارشناسی.xlsx':
            df = pd.read_excel(f)
            merge_excels.append(df)

        if f == r'C:\ezhar-temp\1392\imp_parvand\صد و دویست پرونده انتخابی دادستانی.xlsx':
            df = pd.read_excel(f)
            merge_excels.append(df)
        if f == r'C:\ezhar-temp\%s\%s\Excel.xlsx' % (year, report_type):
            df = pd.read_excel(f)

            if (check_if_col_exists(df, "منبع مالیاتی") == False):
                df.insert(11,
                          column='منبع مالیاتی',
                          value='مالیات بر درآمد شرکت ها')
            merge_excels.append(df)

        if f == r'C:\ezhar-temp\%s\%s\Excel(1).xlsx' % (year, report_type):
            df = pd.read_excel(f)
            if (check_if_col_exists(df, "منبع مالیاتی") == False):
                df.insert(11,
                          column='منبع مالیاتی',
                          value='مالیات بر درآمد مشاغل')
            merge_excels.append(df)

        if f == r'C:\ezhar-temp\%s\%s\Excel(2).xlsx' % (year, report_type):
            df = pd.read_excel(f)
            if (check_if_col_exists(df, "منبع مالیاتی") == False):
                df.insert(11,
                          column='منبع مالیاتی',
                          value='مالیات بر ارزش افزوده')
            merge_excels.append(df)

        if f in [r'C:\ezhar-temp\%s\%s\جزئیات اعتراضات و شکایات.html' % (year, report_type),
                 r'C:\ezhar-temp\%s\%s\گزارش شکايات در جريان دادرسی در هيات بدوی، تجديد نظر و هم عرض.html' % (
                     year, report_type),
                 r'C:\ezhar-temp\%s\%s\جستجو.html' % (year, report_type)]:

            df = pd.read_html(f, flavor="html5lib")[0]
            merge_excels.append(df)

    # Merge all dataframes into one
    final_df_all_fine_grained = pd.concat(merge_excels)
    final_df_all_fine_grained.fillna(0, inplace=True)
    final_df_all_fine_grained.replace('nan', '0', inplace=True)
    final_df_all_fine_grained.replace('None', '0', inplace=True)
    final_df_all_fine_grained.replace('NaN', '0', inplace=True)
    final_df_all_fine_grained.replace('NaT', '0', inplace=True)
    # Find columns with dtype float
    float_columns = final_df_all_fine_grained.select_dtypes(
        include='float64').columns
    # Convert float columns to int
    final_df_all_fine_grained[float_columns] = final_df_all_fine_grained[float_columns].apply(
        np.int64)

    # Clean the Dataframe
    final_df_all_fine_grained = final_df_all_fine_grained.fillna(value=0)
    final_df_all_fine_grained = final_df_all_fine_grained.astype('str')
    final_df_all_fine_grained.replace('0', '', inplace=True)
    final_df_all_fine_grained.replace(0, '', inplace=True)

    final_df_all_fine_grained['تاریخ بروزرسانی'] = get_update_date()

    return final_df_all_fine_grained.values.tolist(
    ), final_df_all_fine_grained.columns


@time_it(log=False)
def connect_to_sql(sql_query,
                   sql_con=get_sql_con(),  # SQL connection string, default value obtained from get_sql_con()
                   # DataFrame values for executemany (if applicable)
                   df_values=None,
                   read_from_sql=False,    # Flag indicating whether to read data from SQL
                   # Connection type (not used in the provided code)
                   connect_type=None,
                   return_df=False,        # Flag indicating whether to return a DataFrame
                   # Chunk size for reading data from SQL (if applicable)
                   chunk_size=None,
                   # Format for returning DataFrame ('json', 'dict', or 'dataframe')
                   return_df_as='dataframe',
                   num_runs=12,            # Maximum number of retry attempts
                   # Additional positional arguments (not used in the provided code)
                   *args,
                   **kwargs):              # Additional keyword arguments (not used in the provided code)

    global n_retries  # Global variable to track the number of retry attempts

    def retry():
        # Establish a connection to SQL Server
        cnxn = pyodbc.connect(sql_con)
        cursor = cnxn.cursor()

        if read_from_sql:
            # Read data from SQL if the flag is set
            df = pd.read_sql(sql_query, cnxn, chunksize=chunk_size)
            return df

        if df_values is None:
            # Execute a single SQL query
            cursor.execute(sql_query)
            cursor.execute('commit')
        else:
            # Execute multiple SQL queries with DataFrame values
            cursor.executemany(sql_query, df_values)
            cursor.execute('commit')

        # Close the SQL connection
        cnxn.close()

    try:
        if return_df:
            # If return_df is True, attempt to execute and return results
            x = retry()

            if return_df_as == 'json':
                # Convert DataFrame to JSON format if specified
                x = x.to_json(orient='records', force_ascii=False)[
                    1:-1].replace('},{', '} {')
            elif return_df_as == 'dict':
                # Convert DataFrame to dictionary format if specified
                x = x.to_dict()

            return x

        else:
            # If return_df is False, only attempt to execute without returning results
            retry()

    except Exception as e:
        # Exception handling for potential errors
        if n_retries < num_runs:
            # Retry if the maximum number of attempts has not been reached
            n_retries += 1
            print(e)
            print('trying again')
            time.sleep(4)
            retry()


class Login:

    def __init__(self, pathsave):
        self.pathsave = pathsave
        fp = set_gecko_prefs(pathsave)
        self.driver = webdriver.Firefox(fp, executable_path=geck_location())
        self.driver.window_handles
        self.driver.switch_to.window(self.driver.window_handles[0])

    def __call__(self):
        return self.driver

    def close(self):
        self.driver.close()


def login_186(driver):
    driver.get("http://govahi186.tax.gov.ir/Login.aspx")
    driver.implicitly_wait(5)

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "MainContent_txtUserName")))
    element.send_keys("1757400389")

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "MainContent_txtPassword")))
    element.send_keys("14579Ali.")

    element = driver.find_element(By.ID, "lblUser")

    try:
        while (element.is_displayed() == False):
            print("waiting for the login to be completed")

    except Exception as e:
        print(e)
    # element = driver.find_element(By.ID, "lblUser")

        return driver
    return driver


def login_tgju(driver):
    driver.get("https://www.tgju.org/")
    driver.implicitly_wait(20)

    return driver


@log_the_func('none')
def login_arzeshafzoodeh(driver, *args, **kwargs):
    success = False
    driver.get("http://10.2.16.131/frmManagerLogin2.aspx")
    driver.implicitly_wait(20)

    while (success == False):
        txtUserName = driver.find_element(By.NAME,
                                          'txtusername').send_keys('959396')
        txtPassword = driver.find_element(By.NAME,
                                          'txtpassword').send_keys('62253LBG')
        time.sleep(8)
        try:
            driver.find_element(By.NAME, 'btnlogin').click()
            try:
                WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(
                        (By.ID,
                         'lblmsg')))
                success = False

            except Exception as e:
                # print('Captcha is wrong')
                success = True
        except Exception:
            success = False

    return driver


def login_soratmoamelat(driver):
    driver.get("http://ittms.tax.gov.ir/")
    driver.implicitly_wait(20)
    txtUserName = driver.find_element(By.ID,
                                      'username').send_keys('1756914443')
    txtPassword = driver.find_element(By.ID,
                                      'Password').send_keys('1756914443')
    time.sleep(10)
    driver.find_element(By.CLASS_NAME, 'button').click()
    time.sleep(1)

    return driver


def login_chargoon(driver=None, info={},
                   user_name='1870051041', password='A3233404'):
    driver.get(
        "http://chargoon-khoozestan:8090/UserLogin.Dap?logout=1&rnd=kqcuhyleisodrvyvhxkhjnlbdjxtjbdu")
    time.sleep(3)
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.ID, "txtClientUsername"))).send_keys(user_name)
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.ID, "txtClientPassword"))).send_keys(password)
    while (True):
        try:
            WebDriverWait(driver, 1).until(
                EC.presence_of_all_elements_located((By.ID, "txtClientPassword")))
            print("Waiting for user to login...")
        except:
            time.sleep(10)
            break

    return driver, info


def login_vosolejra(driver=None, info={},
                    user_name='1910125563', password='1910125563'):
    success = False
    driver.get("http://ve.tax.gov.ir/forms/frmVosool.aspx")
    driver.implicitly_wait(5)
    # Create wait object
    wait = WebDriverWait(driver, 10)

    # Input username
    username_field = wait.until(
        EC.presence_of_element_located((By.ID, "txtUser")))
    username_field.send_keys("1910125563")

    # Input password
    password_field = wait.until(
        EC.presence_of_element_located((By.ID, "txtPass")))
    password_field.send_keys("1910125563")

    # Click login button
    login_button = wait.until(EC.element_to_be_clickable((By.ID, "btnLogin")))
    login_button.click()

    return driver, info


def handle_pred_captcha(driver, file_path, model):
    """
    Predict captcha text and fill the captcha input.
    """
    pred_results = predict_captcha(file_path, model)

    captcha_input = driver.find_element(By.ID, 'CaptchaCode')
    captcha_input.clear()
    captcha_input.send_keys(pred_results[0])

    time.sleep(0.5)


# ==============================
# Environment configuration
# ==============================
IMG_BASE_DIR = os.getenv(
    "CAPTCHA_IMG_DIR",
    r'E:\automating_reports_V2\saved_dir\codeghtesadi\img_files'
)

MODEL_DIR = os.getenv(
    "CAPTCHA_MODEL_DIR",
    r'E:\automating_reports_V2\automation\captcha_pred\saved_models\captcha_model'
)


# ==============================
# Captcha helpers
# ==============================

def handle_pred_captcha(driver, file_path, model):
    pred_results = predict_captcha(file_path, model)

    captcha_input = driver.find_element(By.ID, 'CaptchaCode')
    captcha_input.clear()
    captcha_input.send_keys(pred_results[0])

    time.sleep(0.5)


def handle_captcha_data_gathering(driver, file_path):
    os.makedirs(file_path, exist_ok=True)

    temp_img_path = os.path.join(file_path, 'logo.png')

    with open(temp_img_path, 'wb') as f:
        f.write(
            driver.find_element(By.ID, 'img-captcha')
            .screenshot_as_png
        )

    accepted = False
    while not accepted:
        accepted, file_name = record_keys(until='esc')

    final_path = os.path.join(file_path, f"{file_name}.png")
    os.rename(temp_img_path, final_path)

    driver.refresh()


# ==============================
# Login helpers
# ==============================

def fill_credentials(driver, user_name, password):
    driver.find_element(By.ID, 'username').clear()
    driver.find_element(By.ID, 'username').send_keys(user_name)

    driver.find_element(By.ID, 'Password').clear()
    driver.find_element(By.ID, 'Password').send_keys(password)


def wait_for_4_char_captcha(driver, wait):
    def _has_4_chars(driver):
        value = driver.find_element(
            By.ID, "CaptchaCode").get_attribute("value")
        return len(value) == 4

    wait.until(_has_4_chars)


def submit_login_form(driver, wait):
    captcha_input = wait.until(
        EC.element_to_be_clickable((By.ID, "CaptchaCode"))
    )
    captcha_input.click()

    time.sleep(2)
    wait_for_4_char_captcha(driver, wait)

    submit_label = wait.until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "label.button[for='submit']")
        )
    )
    submit_label.click()


# ==============================
# Main login function
# ==============================

def login_codeghtesadi(
    driver,
    data_gathering=False,
    pred_captcha=False,
    info=None,
    user_name='1756914443',
    password='14579Ali.@'
):
    if info is None:
        info = {}

    LOGIN_PAGE = "https://login.tax.gov.ir/Page/Index"
    REDIRECT_PAGE = "https://login.tax.gov.ir/"

    # 🔒 Guard: login only once per browser session
    if info.get("login_performed"):
        print("ℹ️ Login already handled in this session")
        return driver, info

    # ---------- helpers ----------
    def is_logged_in(wait=False):
        if wait:
            WebDriverWait(driver, 300).until(
                EC.url_matches(LOGIN_PAGE)
            )
        return driver.current_url.startswith(LOGIN_PAGE)

    def is_login_required():
        return driver.current_url.startswith(REDIRECT_PAGE)

    # ---------- open ----------
    driver.get(LOGIN_PAGE)
    time.sleep(2)

    # ✅ session still valid
    if is_logged_in():
        print("✅ Already logged in – using stored session")
        info["login_performed"] = True
        return driver, info

    # ❌ session expired
    if not is_login_required():
        print("⚠️ Unknown login state:", driver.current_url)
        return driver, info

    print("🔐 Session expired – logging in once")

    # ---------- captcha setup ----------
    file_path = os.path.join(
        IMG_BASE_DIR,
        'training' if data_gathering else 'production'
    )

    model = None
    if pred_captcha and not data_gathering:
        model = tf.keras.models.load_model(MODEL_DIR)

    # ---------- credentials ----------
    fill_credentials(driver, user_name, password)

    # ---------- captcha loop ----------
    wait = WebDriverWait(driver, 60)

    while True:
        try:
            if data_gathering:
                handle_captcha_data_gathering(driver, file_path)
                continue

            if pred_captcha:
                handle_pred_captcha(driver, file_path, model)

            submit_login_form(driver, wait)

            if is_logged_in(wait=True):
                print("✅ Login successful – session restored")
                info["login_performed"] = True
                break

        except Exception as e:
            print("⚠️ Login attempt failed, retrying:", e)
            time.sleep(1)

    return driver, info


@wrap_it_with_params(20, 60, True, False, False, False)
def login_mostaghelat(driver, info):
    driver.get("http://most.tax.gov.ir/")
    driver.implicitly_wait(20)
    txtUserName = driver.find_element(By.NAME,
                                      'Txt_username').send_keys('1930841086')
    txtPassword = driver.find_element(By.NAME,
                                      'Txt_Password').send_keys('193084193084')
    time.sleep(0.5)
    driver.find_element(By.NAME, 'login_btn').click()

    return driver, info


def login_hoghogh(driver):
    driver.get("http://10.2.12.234:6399/")
    driver.implicitly_wait(20)

    driver.find_element(By.XPATH,
                        '/html/body/div[1]/div/div[1]/div/div/div/div/div/div/div/form/div[1]/div[1]/label/input[3]').click()
    txtUserName = driver.find_element(By.NAME,
                                      'UserName').send_keys('1756914443')
    txtPassword = driver.find_element(By.NAME,
                                      'Password').send_keys('14579Ali.@')
    time.sleep(0.5)
    driver.find_element(By.ID, 'BtnLogin').click()

    return driver


def check_if_login_iris_user_pass_inserted(driver, creds):
    if (creds['username'] == WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable(
            (By.NAME,
                'userName'))).get_attribute('value') and str(creds['pass']) == WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable(
            (By.NAME,
                'password'))).get_attribute('value')):
        return driver
    else:
        raise Exception


# @wrap_a_wrapper
@wrap_it_with_params(20, 60, True, False, False, False)
def insert_login_iris_user_pass(driver, creds, info):
    WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable(
            (By.NAME,
                'userName'))).clear()
    WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable(
            (By.NAME,
                'userName'))).send_keys(creds['username'])

    # WebDriverWait(driver, 2).until(
    #     EC.element_to_be_clickable(
    #         (By.NAME,
    #             'userName'))).send_keys('')

    WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable(
            (By.NAME,
                'password'))).clear()

    WebDriverWait(driver, 2).until(
        EC.element_to_be_clickable(
            (By.NAME,
                'password'))).send_keys(creds['pass'])
    driver = check_if_login_iris_user_pass_inserted(driver, creds)

    return driver, info


# @wrap_a_wrapper
@wrap_it_with_params(20, 60, True, False, False, False)
def login_iris(driver, creds=None, info={'success': True}):
    driver.get("https://its.tax.gov.ir/flexform2/logineris/form2login.jsp")

    time.sleep(4)

    driver, info = insert_login_iris_user_pass(
        driver=driver, creds=creds, info=info)

    time.sleep(2)

    WebDriverWait(driver, 3).until(
        EC.element_to_be_clickable(
            (By.ID,
                'ok_but'))).click()
    time.sleep(0.5)

    try:
        alert_obj = driver.switch_to.alert
        alert_obj.accept()
        error = True
        time.sleep(5)
    except:
        error = False

    if error:
        raise Exception

    driver, info = find_obj_login(driver=driver, info=info)

    return driver, info


# @wrap_a_wrapper
@wrap_it_with_params(20, 60, True, True, False, False)
def login_list_hoghogh(driver, info={}, creds={'username': '1756914443',
                                               'password': '1756914443',
                                               'username_modi': '10101862318008'}):
    driver.get("http://salary.tax.gov.ir/Account/LogOnArshad")

    txtUserName = driver.find_element(By.ID,
                                      'UserNameArshad').send_keys(creds['username'])
    txtPassword = driver.find_element(By.ID,
                                      'PasswordArshad').send_keys(creds['password'])
    txtPassword = driver.find_element(By.ID,
                                      'UserNameMoadi').send_keys(creds['username_modi'])

    driver.find_element(
        By.XPATH, '/html/body/div[2]/div[2]/div/fieldset[1]/form/div[7]/p/input').click()

    return driver, info


@wrap_it_with_params(20, 60, True, True, False, False)
def login_nezam_mohandesi(driver, info={}, creds={'username': '1756914443',
                                                  'password': '1756914443',
                                                  'username_modi': '10101862318008'}):
    driver.get("https://reports.khzceo.ir/m5_done.aspx")

    return driver, info


@wrap_it_with_params(15, 60, True, False, False, False)
def login_sanim(driver, info):
    driver.get("https://mgmt.tax.gov.ir/ords/f?p=100:101:16540338045165:::::")

    txtUserName = driver.find_element(By.ID,
                                      'P101_USERNAME').send_keys('1910125563')
    txtPassword = driver.find_element(By.ID,
                                      'P101_PASSWORD').send_keys('123456')

    try:
        while (driver.find_element(By.XPATH,
                                   '/html/body/form/div/main/div/div/div/div/div[2]/div[4]/\
                                          div[2]/div/div/center/img[2]').is_displayed()):
            time.sleep(1)

            driver.find_element(By.ID, 'B1700889564218640').click()
    except:
        info['cur_instance'] = driver.current_url.split(':')[-6]

    return driver, info


def get_tblreports_date(tblnames,
                        years=['1395', '1396', '1397', '1398', '1399']):

    lst = []

    for year in years:

        for tblname in tblnames:

            sql_query = """
            
            select distinct [تاریخ بروزرسانی] from [dbo].[%s%s] 
            
            """ % (tblname, year)

            date = connect_to_sql(sql_query,
                                  sql_con=get_sql_con(server='10.52.0.114'),
                                  read_from_sql=True,
                                  connect_type='',
                                  return_df=True)

            if not (date.empty):
                date = date.values.tolist()
                lst.append((tblname, year, date[0][0]))

    return lst


@measure_time
def drop_into_db(table_name=None,
                 columns=None,
                 values=None,
                 append_to_prev=False,
                 db_name='TestDb',
                 del_tbl=None,
                 create_tbl=None,
                 sql_query=None,
                 drop=True):
    # If not appending to previous data
    if not append_to_prev:
        # Deleting the previous table if specified
        def del_tbl_func():
            delete_table = sql_delete(table_name)
            connect_to_sql(sql_query=delete_table,
                           sql_con=get_sql_con(database=db_name),
                           connect_type='dropping sql table')

        # If no specific instructions are given for deleting the table
        if del_tbl is None:
            del_tbl_func()

        # Creating a new table if specified or if default behavior is to create
        def create_tbl_func():
            sql_create_table = create_sql_table(table_name, columns)
            connect_to_sql(sql_create_table,
                           sql_con=get_sql_con(database=db_name),
                           connect_type='creating sql table')

        # If no specific instructions are given for creating the table or if default behavior is to create
        if (create_tbl is None) or (create_tbl == 'yes'):
            create_tbl_func()

    # Inserting data into the table
    if drop:
        # If no specific SQL query is provided, use the default insert query
        if sql_query is None:
            sql_query = insert_into(table_name, columns)

        # Connect to SQL and execute the insert query
        connect_to_sql(sql_query,
                       sql_con=get_sql_con(database=db_name),
                       df_values=values,
                       connect_type='inserting into sql table')


def get_edare_shahr():
    sql_query_edare_shahr = "SELECT [city],[edare],[IrisCode] FROM [tax].[dbo].[tblEdareShahr]"
    df_edare_shahr = connect_to_sql(sql_query_edare_shahr,
                                    sql_con=get_sql_con(database='tax'),
                                    read_from_sql=True,
                                    connect_type='read from tblEdateShahr',
                                    return_df=True)
    return df_edare_shahr


def find_edares(df=None, colname_tomerge='واحد مالیاتی', persian_name=None, table_name='None', report_typ=None):
    df_edare_shahr = get_edare_shahr()

    if df is None:
        sql_query_most = "SELECT * FROM %s" % table_name
        df = connect_to_sql(sql_query_most,
                            sql_con=get_sql_con(database='testdbV2'),
                            read_from_sql=True,
                            connect_type='read from tblMost',
                            return_df=True)
        df = df.drop(columns=['ID'], axis=1)

    df_merged_1 = df.merge(df_edare_shahr,
                           how='inner',
                           left_on=df[colname_tomerge].astype(
                               'str').str.slice(0, 4),
                           right_on='edare')

    df_merged_2 = df.merge(df_edare_shahr,
                           how='inner',
                           left_on=df[colname_tomerge].astype(
                               'str').str.slice(0, 5),
                           right_on=df_edare_shahr['edare'].str.slice(
                               0, 5)).drop(['key_0'], axis=1)

    df_f = pd.concat([df_merged_1, df_merged_2])
    if report_typ is not None:
        df_f.drop_duplicates(subset=[report_typ], keep=False, inplace=True)

    dff_final = pd.concat([df_f, df_merged_2])

    dff_final.rename(columns={
        'edare': 'کد اداره',
        'city': 'شهرستان'
    },
        inplace=True)

    return dff_final


def process_mostaghelat(date=140106,
                        report_type='Tashkhis',
                        persian_name='تشخیص',
                        sodor='صدور',
                        msg='تشخیص ابلاغ نشده',
                        table_name=None,
                        report_typ='شماره برگ تشخیص',
                        drop_to_sql=True):

    dff_final = find_edares(table_name=table_name, report_typ=report_typ)

    dff_final['ماه %s' % sodor] = dff_final['تاریخ %s' % sodor].str.replace(
        '/', '').str.slice(0, 6).astype('int64')
    if date is not None:
        dff_final = dff_final.loc[dff_final['ماه %s' % sodor] < 140106]
    dff_final.rename(columns={
        'edare': 'کد اداره',
        'city': 'شهرستان'
    },
        inplace=True)
    dff_final_agg = dff_final.groupby(
        ['کد اداره', 'شهرستان', 'تاریخ بروزرسانی']).size().reset_index()
    dff_final_agg.rename(
        columns={0: 'تعداد %s مستغلات' % msg}, inplace=True)
    # drop_into_db(table_name, columns, values)
    if drop_to_sql:
        drop_into_db(table_name,
                     dff_final.columns.tolist(),
                     dff_final.values.tolist(),
                     append_to_prev=False,
                     db_name='testdbV2')
    return dff_final, dff_final_agg


def final_most(date=140106):
    tashkhis, tashkhis_agg = process_mostaghelat(date=date,
                                                 table_name='tblTashkhisMost', report_typ='شماره برگ تشخیص')
    ghatee, ghatee_agg = process_mostaghelat(date=date,
                                             report_type='Ghatee',
                                             persian_name='قطعی',
                                             msg='قطعی ابلاغ نشده',
                                             report_typ='شماره برگ قطعی',
                                             table_name='tblGhateeMost')
    amade_ghatee, amade_ghatee_agg = process_mostaghelat(
        date=None,
        report_type='AmadeGhatee',
        persian_name='تشخیص',
        sodor='ابلاغ',
        msg='آماده قطعی',
        report_typ='شماره برگ تشخیص',
        table_name='tblAmadeGhateeMost')

    lst_agg = [tashkhis_agg, ghatee_agg, amade_ghatee_agg]

    merged_agg = ft.df_to_excelsheete(
        lambda left, right: pd.merge(
            left, right, how='outer', on='کد اداره'),
        lst_agg)

    merged_agg['شهرستان'].fillna(
        merged_agg['شهرستان_y'], inplace=True)
    merged_agg['شهرستان'].fillna(
        merged_agg['شهرستان_x'], inplace=True)
    merged_agg['تاریخ بروزرسانی'].fillna(merged_agg['تاریخ بروزرسانی_x'],
                                         inplace=True)
    merged_agg['تاریخ بروزرسانی'].fillna(merged_agg['تاریخ بروزرسانی_y'],
                                         inplace=True)
    merged_agg['شهرستان'].fillna(
        merged_agg['شهرستان_x'], inplace=True)
    merged_agg.fillna(0, inplace=True)

    selected_columns = [
        'کد اداره', 'تعداد تشخیص ابلاغ نشده مستغلات',
        'تعداد قطعی ابلاغ نشده مستغلات', 'شهرستان', 'تاریخ بروزرسانی',
        'تعداد آماده قطعی مستغلات'
    ]

    merged_agg = merged_agg[selected_columns]
    merged_agg = merged_agg.rename(
        columns={'کد اداره': 'نام اداره سنتی'})
    merged_agg['نام اداره سنتی'] = merged_agg['نام اداره سنتی'].str.slice(
        0, 5)
    merged_agg = merged_agg.iloc[:, [0, 3, 1, 2, 5, 4]]

    return tashkhis, ghatee, amade_ghatee, merged_agg


def rename_duplicate_columns(df):

    # Create a Series containing the original column names
    cols = pd.Series(df.columns)

    # Iterate through duplicate column names
    for dup in df.columns[df.columns.duplicated(keep=False)]:
        # Update column names with a numerical suffix for duplicates
        cols[df.columns.get_loc(dup)] = ([
            dup + '.' + str(d_idx) if d_idx != 0 else dup
            for d_idx in range(df.columns.get_loc(dup).sum())
        ])

    # Assign the modified column names back to the DataFrame
    df.columns = cols

    return df


def georgian_to_persian(row, type_of='g', delimeter='-', complete_date=False):

    none_date = None  # Define the value to return for undefined dates
    # List of placeholders for undefined dates
    lst_none = ['None', 'NaT', 'nan']

    if type_of == 't':
        if row in lst_none:
            return none_date
        else:
            x = int(row[:6])
            return x
    elif type_of == 'g':
        if row in lst_none:
            return none_date
        else:
            lst = row.split(delimeter)
            # Assuming there's a function 'get_update_date' to convert the Georgian date to Persian
            date = get_update_date(lst)

            if complete_date:
                return date

            return int(date[:8])


def is_date(str, fuzzy=False):

    lst_none = ['None', 'NaT', 'nan']

    try:
        # Check if the input string is in the list of known non-date values
        if str in lst_none:
            return str

        # Extract the first 10 characters (representing the date part) from the input string
        str = str[:10]

        # Attempt to parse the date using the `parse` function
        parse(str, fuzzy=fuzzy)

        # Convert the parsed date to a Persian date using the `georgian_to_persian` function
        str = georgian_to_persian(str, complete_date=True)

        # Return the Persian date string
        return str

    except ValueError:
        # If an error occurs during parsing, return the original input string
        return str


def is_int(str):

    try:
        # Try converting the string to a float
        str = float(str)

        # Check if the result is a float
        if isinstance(str, float):
            # If it's a float, convert it to an integer
            str = int(str)

        # Return the result (either the integer or the original string)
        return str

    except Exception as e:
        # If an exception occurs during the conversion, catch and print the exception (optional)
        # print(e)

        # Return the original string if conversion fails
        return str


def insert_into_tbl(sql_query, tbl_name, convert_to_str=False):

    # Retrieve data from SQL
    df = connect_to_sql(sql_query,
                        read_from_sql=True,
                        return_df=True,
                        connect_type='Read from table')

    # Optionally convert columns to string type
    if convert_to_str:
        df = df.astype('str')

    # Drop 'ID' column if present
    cols = df.columns.tolist()
    if 'ID' in cols:
        df.drop(['ID'], axis=1, inplace=True)

    # Ensure all columns contain integer values by applying the 'is_int' function
    cols = df.columns.tolist()
    for item in cols:
        df[item] = df[item].apply(lambda x: is_int(x))

    # Add 'تاریخ بروزرسانی' column with the current update date
    df['تاریخ بروزرسانی'] = get_update_date()

    # Optionally convert columns to string type again
    if convert_to_str:
        df = df.astype('str')

    # Insert data into the specified table
    drop_into_db(table_name=tbl_name,
                 columns=df.columns.tolist(),
                 values=df.values.tolist(),
                 db_name='testdbV2')


def return_start_end(s=0, end=1000000000000000000, inc=10000, alpha_inc=99.1):

    # Initialize an empty list to store tuples
    lst = []

    # Initialize the step with the initial increment
    step = inc

    # Loop until the step exceeds the specified end point
    while step <= end:
        # Append tuple (start, end) to the list
        lst.append((int(s), int(step)))

        # Update starting point and increment values
        s = step + 1
        step += inc
        inc *= alpha_inc

    # Note: The following two lines define additional static points (start_p and end_p)
    #       but are not used in the function's logic. They can be safely removed.
    # start_p = 0
    # end_p = 10000

    # Return the list of tuples representing start and end points
    return lst


def list_files(path, extension):

    file_list = glob.glob(path + "/*" + extension)
    return file_list


def get_files_sizes(path, postfixes):
    try:
        # Check for the presence of temporary download files
        file_paths = [glob.glob(path + "/*" + '.%s.part' % postfix)
                      for postfix in postfixes]
        return os.path.getsize(file_paths[1][0])
    except FileNotFoundError:
        return 0


def wait_for_download_to_finish(path, postfixes=['xls', 'html'], sleep_time=6):

    # Sleep to allow for initial file creation
    time.sleep(sleep_time)

    # Variables to track download status
    file_exists = False
    i = 0
    file_size = 0
    while not file_exists:
        time.sleep(sleep_time)
        i += sleep_time

        # Print a message every 30 seconds to indicate ongoing download
        if i % (sleep_time * 2) == 0:

            print(
                f'Waiting for the download to finish, time elapsed is {i} seconds')

        time.sleep(6)
        # Check for the presence of temporary download files
        file_lists = [glob.glob(path + "/*" + '.%s.part' % postfix)
                      for postfix in postfixes]

        file_exists = not any(len(f) != 0 for f in file_lists)
        if not file_exists:
            current_file_size = get_files_sizes(path, postfixes)
            if current_file_size > file_size:
                file_size = current_file_size
                continue
            else:
                return 0
    return 1


def extract_nums(text, expression=r'\d+'):
    # Extract numbers using regex
    numbers = re.findall(expression, text)
    return numbers


def unzip_files(dir, remove=True):
    # Step 1: Find all subdirectories in the specified directory
    dirs = [it.path for it in os.scandir(dir) if it.is_dir()]

    # Step 2: Find subdirectories within each subdirectory (nested directories)
    final_dirs = dirs
    for item in dirs:
        final_dirs.extend([it.path for it in os.scandir(item) if it.is_dir()])

    # Step 3: Iterate through all found directories
    for item in final_dirs:
        # Step 4: Find all zip files in the current directory
        files = list_files(item, 'zip')

        # Step 5: If zip files are found, unzip each file
        if len(files) > 0:
            for file in files:
                with zipfile.ZipFile(file, 'r') as zipObj:
                    # Extract all contents of the zip file into the current directory
                    zipObj.extractall(item)

            # Step 6: If 'remove' is True, delete the original zip files
            if remove:
                for file in files:
                    os.remove(file)

# Additional function to list files with a specific extension in a directory


def list_files(directory, extension):
    return [it.path for it in os.scandir(directory) if it.is_file() and it.name.endswith('.' + extension)]

# Example usage:
# Specify the directory path where you want to unzip files
# unzip_files('/path/to/directory', remove=True)


def zipdir(path, ziph):

    # Iterate over all the directories and files in the given path
    for (root, dirs, files) in os.walk(path):
        print(root)  # Print the current directory being processed
        for file in files:
            # Write each file to the zip file with a relative path
            ziph.write(
                os.path.join(root, file),
                os.path.relpath(os.path.join(root, file),
                                os.path.join(path, '..'))
            )


def get_file_size(filename):
    return os.stat(filename)


@log_the_func('none')
def move_files(srcs, dsts, *args, **kwargs):
    # Iterate over pairs of source and destination paths
    for src, dst in zip(srcs, dsts):
        # Move the file from source to destination
        shutil.move(src, dst)


def df_to_excelsheet(saved_path, dict_df, index, *args, **kwargs):

    # Use pd.ExcelWriter to efficiently write multiple DataFrames to an Excel file
    with pd.ExcelWriter(saved_path) as writer:
        for key, dataframe in dict_df.items():
            # Handle tuple keys for composite sheet names
            if isinstance(key, tuple):
                sheet_name = ''
                for item in key:
                    # Remove specified names from the sheet name
                    if 'names' in kwargs:
                        for name in kwargs['names']:
                            if name in item:
                                item = item.replace(name, '')
                    sheet_name += item + '-'
            else:
                # Convert non-tuple keys to string for sheet name
                sheet_name = str(key)

            # Write DataFrame to Excel sheet
            dataframe.to_excel(writer, sheet_name=sheet_name, index=index)


def df_to_excelworkbooks(df, g_cols, saving_path=None,
                         zipped_output={'zipped': True,
                                        'dir': r'E:\automating_reports_V2\saved_dir\arzeshafzoodeh_sonati\test\final'},
                         *args, **kwargs):

    # Set default saving path if not provided
    if saving_path is None:
        saving_path = r'E:\automating_reports_V2\saved_dir\arzeshafzoodeh_sonati\test\final\test'

    # Group the DataFrame based on specified columns
    grouped_df = df.groupby(g_cols)

    # Iterate through each group and save as separate Excel workbook
    for key, item in grouped_df:
        saved_dir = os.path.join(saving_path, key)
        if not os.path.exists(saved_dir):
            os.makedirs(saved_dir)
        name = key + '.xlsx'
        item.to_excel(os.path.join(saved_dir, name))

    # Check if zipping is enabled
    if zipped_output['zipped']:
        # Create the directory for zipped output if it doesn't exist
        if not os.path.exists(zipped_output['dir']):
            os.makedirs(zipped_output['dir'])
        # Create a zip file and add the contents of the saving path
        file_name = os.path.join(zipped_output['dir'], 'final.zip')
        with zipfile.ZipFile(file_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipdir(saving_path, zipf)

# Helper methods for predicting Captcha


def get_lst_images(data_dir):
    # Get list of all the images
    images = sorted(list(map(str, list(data_dir.glob("*.png")))))
    labels = [img.split(os.path.sep)[-1].split(".png")[0] for img in images]
    characters = set(char for label in labels for char in label)
    characters = sorted(list(characters))

    print("Number of images found: ", len(images))
    print("Number of labels found: ", len(labels))
    print("Number of unique characters: ", len(characters))
    print("Characters present: ", characters)

    return images, labels, characters


def decode_batch_predictions(pred, max_length, num_to_char):
    input_len = np.ones(pred.shape[0]) * pred.shape[1]
    # Use greedy search. For complex tasks, you can use beam search
    results = keras.backend.ctc_decode(pred, input_length=input_len, greedy=True)[0][0][
        :, :max_length
    ]
    # Iterate over the results and get back the text
    output_text = []
    for res in results:
        res = tf.strings.reduce_join(num_to_char(res)).numpy().decode("utf-8")
        output_text.append(res)
    return output_text

#  Let's check results on some validation samples


def validate_results(training=False, validation_dataset=None, prediction_model=None, max_length=4, num_to_char=4):
    for batch in validation_dataset.take(1):
        batch_images = batch["image"]
        batch_labels = batch["label"]

        preds = prediction_model.predict(batch_images)
        pred_texts = decode_batch_predictions(
            preds, max_length=max_length, num_to_char=num_to_char)

        return pred_texts


def predict_captcha(data_dir_prod, model):

    img_width = 200
    img_height = 50

    x_images, labels, characters = get_lst_images(Path(data_dir_prod))

    max_length = max([len(label) for label in labels])

    # Mapping characters to integers
    char_to_num = layers.StringLookup(
        vocabulary=list(characters), mask_token=None
    )

    # Mapping integers back to original characters
    num_to_char = layers.StringLookup(
        vocabulary=char_to_num.get_vocabulary(), mask_token=None, invert=True
    )

    def encode_single_sample(img_path, label):

        # 1. Read image
        img = tf.io.read_file(img_path)
        # 2. Decode and convert to grayscale
        img = tf.io.decode_png(img, channels=1)
        # 3. Convert to float32 in [0, 1] range
        img = tf.image.convert_image_dtype(img, tf.float32)
        # 4. Resize to the desired size
        img = tf.image.resize(img, [img_height, img_width])
        # 5. Transpose the image because we want the time
        # dimension to correspond to the width of the image.
        img = tf.transpose(img, perm=[1, 0, 2])
        # 6. Map the characters in label to numbers
        label = char_to_num(tf.strings.unicode_split(
            label, input_encoding="UTF-8"))
        # 7. Return a dict as our model is expecting two inputs
        return {"image": img, "label": label}

    model_save_dir = r'E:\automating_reports_V2\automation\captcha_pred\saved_models\captcha_model'

    prediction_model = keras.models.Model(
        model.get_layer(name="image").input, model.get_layer(
            name="dense2").output
    )

    validation_dataset = tf.data.Dataset.from_tensor_slices(
        (np.array(x_images), np.array(labels)))

    validation_dataset = (
        validation_dataset.map(
            encode_single_sample, num_parallel_calls=tf.data.AUTOTUNE
        )
        .batch(1)
        .prefetch(buffer_size=tf.data.AUTOTUNE)
    )
    # prediction_model.summary()
    pred_texts = validate_results(
        training=False, validation_dataset=validation_dataset, prediction_model=prediction_model,
        max_length=max_length, num_to_char=num_to_char)

    return pred_texts


def count_num_files(dir):
    """A python function which return the number of files in a directory"""
    count = 0
    for _, _, files in os.walk(dir):
        count += len(files)
    return count


def get_sqltable_colnames(db_name, tbl_name):

    # Construct SQL query to retrieve column names
    sql_query = f"""SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE 
                TABLE_CATALOG = '{db_name}'
            AND TABLE_NAME = '{tbl_name}'
            """

    # Connect to SQL and execute the query
    df_cols = connect_to_sql(
        sql_query, sql_con=get_sql_con(database=db_name), read_from_sql=True, return_df=True)

    # Reshape the result to a flat list of column names
    df_cols = list(np.reshape(df_cols.to_numpy(), (-1, 1)).flatten())

    # Return the list of column names
    return df_cols


def highlight_excel_cells(df, rows_range, cols_range=None, saved_path=None, hex_color='4d4dff'):

    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Write the DataFrame to the Excel sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # If cols_range is not provided, generate default columns (A-Z)
    if cols_range is None:
        max_cols = math.floor((ws.max_column / 26))
        lst_alphabet = list(utils.get_column_letter(i) for i in range(1, 27))

        cols_range = []
        cols_range.extend(lst_alphabet)

        for item in lst_alphabet[:max_cols]:
            for w in lst_alphabet:
                cols_range.append(f'{item}{w}')

    # Highlight specified cells in the given rows and columns with the specified hex color
    for i in range(1, rows_range + 1):  # Note: Adjusted range to include the last row
        for item in cols_range:
            cell = ws['%s%s' % (item, i)]
            # Uncomment the line below to change font color (example: red)
            # cell.font = styles.Font(color="FF0000")
            cell.fill = styles.fills.PatternFill(
                patternType='solid', fgColor=styles.Color(rgb=hex_color))

    # Save the workbook to the specified path, if provided
    if saved_path is not None:
        wb.save(saved_path)


def generate_readme(file_path):
    content = """# Project Title

Brief description of your project.

## Table of Contents

- [Installation](#installation)
- [Usage](#usage)
- [Contributing](#contributing)
- [License](#license)

## Installation

Describe how to install your project.

## Usage

Provide instructions on how to use your project.

## Contributing

Explain how others can contribute to your project.

## License

This project is licensed under the [License Name] - see the [LICENSE.md](LICENSE.md) file for details.

```python
def example_function():
    # Your code here
    pass
"""

    with open(file_path, 'w') as readme_file:
        readme_file.write(content)


# Specify the file path where you want to save the README.md file
readme_file_path = 'README.md'


def get_child_processes(parent_pid):
    child_processes = []

    try:
        process = psutil.Process(parent_pid)
        for child in process.children(recursive=True):
            child_processes.append(child.pid)
    except psutil.NoSuchProcess:
        pass

    return child_processes


def cleanup(pid):
    # Terminate all subprocesses when the application is closed
    # Note: This is a simple example; you may need to adapt it based on your specific subprocess handling
    print('cleaning up')
    procs = get_child_processes(pid)
    print(procs)
    for proc in psutil.process_iter(['pid', 'cmdline']):
        if proc.pid in procs:
            proc.terminate()


def add_days_to_persian_date(date_str, days=1):

    # Convert the string to a Persian date
    persian_date = datetime.strptime(date_str, '%Y/%m/%d')

    # Add one day
    persian_date += timedelta(days=1)

    # Convert the Persian date to a string
    persian_date_str = persian_date.strftime('%Y/%m/%d')
    return persian_date_str


def calculate_days_difference(date1: str, date2: str):

    if len(date1) < 5:
        return 0

    date1 = date1.replace("/", "-")
    date2 = date2.replace("/", "-")
    jalali_date1 = JalaliDate.fromisoformat(date1)
    jalali_date2 = JalaliDate.fromisoformat(date2)

    gregorian_date1 = jalali_date1.to_gregorian()
    gregorian_date2 = jalali_date2.to_gregorian()

    day_difference = (gregorian_date2 - gregorian_date1).days
    return abs(day_difference)


if __name__ == "__main__":
    calculate_days_difference('1403/09/07', get_update_date(delimiter='/'))
